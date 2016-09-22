"""
    Author : Sumit Kumar

    Details: This script parses the ASUP Excel file,
    and creates a new output file with unique customer names
    and other relevant data.

    Required packages: openpyxl
    To Run : python ./ASUP_Excel_Parser.py <PATH TO EXCEL FILE>

    Input : ASUP report file to parse.
    Output : Output.xlsx
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font
import warnings


class CustomerClass(object):
    def __init__(self):
        self.customerName = None
        self.parentName = None
        self.serialNumbers = set()
        self.CinderOnONTAP = 0
        self.ManilaOnONTAP = 0
        self.CinderOnESeries = 0


# Ensure that the input is valid and the file exists.
if not (len(sys.argv) == 2 or len(sys.argv) == 3):
    print "Error: Invalid argument length."
    print "To Run : python ./ASUP_Excel_Parser.py <PATH TO EXCEL FILE> " \
          "<Optional - OUTPUT FILE NAME>"
    exit()

filePath = sys.argv[1]
outputFilePath = None

if len(sys.argv) == 3:
    outputFilePath = sys.argv[2]

if os.path.exists(filePath):
    print "Excel file found."
    print "Loading..."
else:
    print "Excel file not found"
    exit()

inputExcel = None

try:
    warnings.simplefilter("ignore")
    inputExcel = openpyxl.load_workbook(filePath)
except:
    print "Error loading Excel file."
    exit()

# sheet = inputExcel.get_sheet_by_name(inputExcel.get_sheet_names()[0])

sheet = inputExcel.active
max_row = sheet.max_row
max_column = sheet.max_column
customerDict = dict()


for i in range(2, max_row + 1):

    customerName = sheet.cell(row=i, column=27).value
    serialNumber = sheet.cell(row=i, column=2).value
    sysModel = sheet.cell(row=i, column=4).value.lower()
    subject = sheet.cell(row=i, column=12).value.lower()

    # ToDo: Check if this is the right way to detect E-Series/ONTAP
    usesESeries=0
    usesONTAP=0
    if sysModel == "santricity":
        usesESeries = True
    else:
        usesONTAP = True

    # ToDo: Check if this is the right way to detect Cinder/Manila
    usesCinder = False
    usesManila = False
    if "cinder" in subject:
        usesCinder = True
    elif "manila" in subject:
        usesManila = True

    # Instantiate customerDetails based on whether this
    # customer has been parsed before
    customerDetails = None
    if customerName in customerDict:
        # Customer has been parsed before. Update/Append details.
        customerDetails = customerDict[customerName]

    else:
        # New customer
        customerDetails = CustomerClass()

    # Add/Update customer details

    controllerParsedBefore = False
    if serialNumber in customerDetails.serialNumbers:
        controllerParsedBefore = True

    if usesESeries is True and controllerParsedBefore is False:
        customerDetails.CinderOnESeries += 1
    if usesONTAP is True:
        if usesCinder is True:
            if controllerParsedBefore is False:
                customerDetails.CinderOnONTAP += 1
            elif customerDetails.CinderOnONTAP == 0:
                # Check to not count duplicate reports.
                # Cinder & Manila both can be used on same ONTAP controller.
                customerDetails.CinderOnONTAP += 1
        elif usesManila is True:
            if controllerParsedBefore is False:
                customerDetails.ManilaOnONTAP += 1
            elif customerDetails.ManilaOnONTAP == 0:
                # Check to not count duplicate reports.
                # Cinder & Manila both can be used on same ONTAP controller.
                customerDetails.ManilaOnONTAP += 1

    customerDetails.serialNumbers.add(serialNumber)
    # Add/Update dictionary
    customerDict[customerName] = customerDetails

# Write output data to Excel file.
print "Writing output file..."
outputExcel = openpyxl.Workbook()
newSheet = outputExcel.active

newSheet.cell(row=1, column=1).value = "Customer"
newSheet.cell(row=1, column=2).value = "Controllers"
newSheet.cell(row=1, column=3).value = "Cinder on ONTAP"
newSheet.cell(row=1, column=4).value = "Manila on ONTAP"
newSheet.cell(row=1, column=5).value = "Cinder on E-Series"

# Apply styling
headingFont = Font(size=12, bold=True)
newSheet.cell(row=1, column=1).font = headingFont
newSheet.cell(row=1, column=2).font = headingFont
newSheet.cell(row=1, column=3).font = headingFont
newSheet.cell(row=1, column=4).font = headingFont
newSheet.cell(row=1, column=5).font = headingFont
newSheet.column_dimensions['A'].width = 30
newSheet.column_dimensions['B'].width = 17
newSheet.column_dimensions['C'].width = 17
newSheet.column_dimensions['D'].width = 17
newSheet.column_dimensions['E'].width = 17

# Fill data
rowIndex = 2
for key in customerDict:
    newSheet.cell(row=rowIndex, column=1).value = key
    newSheet.cell(row=rowIndex, column=2).value = customerDict[key].serialNumbers.__len__()
    newSheet.cell(row=rowIndex, column=3).value = customerDict[key].CinderOnONTAP
    newSheet.cell(row=rowIndex, column=4).value = customerDict[key].ManilaOnONTAP
    newSheet.cell(row=rowIndex, column=5).value = customerDict[key].CinderOnESeries

    rowIndex += 1

try:
    if outputFilePath is None:
        outputFilePath = 'Output.xlsx'
    if not outputFilePath.__contains__('.xlsx'):
        outputFilePath += '.xlsx'
    outputExcel.save(outputFilePath)
    print "Success!"
except:
    print "Unable to save file. Please ensure you have sufficient disk space " \
          "and permissions."




