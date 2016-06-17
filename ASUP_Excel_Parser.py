"""
    Author : Sumit Kumar

    Details: This script parses the ASUP Excel file, and creates a new output file with unique customer names and other relevant data.

    Required packages: openpyxl
    To Run : python ./ASUP_Excel_Parser.py <PATH TO EXCEL FILE>

    Input : ASUP report file to parse.
    Output : Output.xlsx
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font
import warnings


class CustomerClass(object):

    def __init__(self):
        self.customerName = None
        self.parentName = None
        self.serialNumbers = set()
        self.usesManila = False
        self.usesCinder = False
        self.usesESeries = False
        self.usesONTAP = False


# Ensure that the input is valid and the file exists.
if len(sys.argv) is not 2:
    print "Error: Invalid argument length."
    print "To Run : python ./ASUP_Excel_Parser.py <PATH TO EXCEL FILE>"
    exit()

filePath = sys.argv[1]

if os.path.exists(filePath):
    print "Excel file found."
    print "Loading..."
else:
    print "Excel file not found"
    exit()

inputExcel = None

try:
    warnings.simplefilter("ignore")
    inputExcel = openpyxl.load_workbook(filePath)
except:
    print "Error loading Excel file."
    exit()

# sheet = inputExcel.get_sheet_by_name(inputExcel.get_sheet_names()[0])

sheet = inputExcel.active
max_row = sheet.max_row
max_column = sheet.max_column
customerDict = dict()


for i in range(2, max_row + 1):

    customerName = sheet.cell(row=i, column=27).value
    serialNumber = sheet.cell(row=i, column=2).value
    sysModel = sheet.cell(row=i, column=4).value.lower()
    subject = sheet.cell(row=i, column=12).value.lower()

    # ToDo: Check if this is the right way to detect E-Series/ONTAP
    usesESeries=False
    usesONTAP=False
    if sysModel == "santricity":
        usesESeries = True
    else:
        usesONTAP = True

    # ToDo: Check if this is the right way to detect Cinder/Manila
    usesCinder = False
    usesManila = False
    if "cinder" in subject:
        usesCinder = True
    elif "manila" in subject:
        usesManila = True

    if customerName in customerDict:
        # Customer has been parsed before. Update/Append details.
        customerDetails = customerDict[customerName]
        customerDetails.serialNumbers.add(serialNumber)

        if customerDetails.usesESeries is False and usesESeries is True:
            customerDetails.usesESeries = True
        if customerDetails.usesONTAP is False and usesONTAP is True:
            customerDetails.usesONTAP = True

        if customerDetails.usesCinder is False and usesCinder is True:
            customerDetails.usesCinder = True
        if customerDetails.usesManila is False and usesManila is True:
            customerDetails.usesManila = True

    else:
        # New customer, add details.
        customerDetails = CustomerClass()
        customerDetails.serialNumbers.add(serialNumber)
        customerDetails.usesESeries = usesESeries
        customerDetails.usesONTAP = usesONTAP
        customerDetails.usesCinder = usesCinder
        customerDetails.usesManila = usesManila

    # Add or update
    customerDict[customerName] = customerDetails

# Write output data to Excel file.
print "Writing output file..."
outputExcel = openpyxl.Workbook()
newSheet = outputExcel.active

newSheet.cell(row=1, column=1).value = "Customer"
newSheet.cell(row=1, column=2).value = "Controllers"
newSheet.cell(row=1, column=3).value = "Cinder"
newSheet.cell(row=1, column=4).value = "Manila"
newSheet.cell(row=1, column=5).value = "ONTAP"
newSheet.cell(row=1, column=6).value = "E-Series"

# Apply styling
headingFont = Font(size=12, bold=True)
newSheet.cell(row=1, column=1).font = headingFont
newSheet.cell(row=1, column=2).font = headingFont
newSheet.cell(row=1, column=3).font = headingFont
newSheet.cell(row=1, column=4).font = headingFont
newSheet.cell(row=1, column=5).font = headingFont
newSheet.cell(row=1, column=6).font = headingFont
newSheet.column_dimensions['A'].width = 30

# Fill data
rowIndex = 2
for key in customerDict:
    newSheet.cell(row=rowIndex, column=1).value = key
    newSheet.cell(row=rowIndex, column=2).value = customerDict[key].serialNumbers.__len__()
    newSheet.cell(row=rowIndex, column=3).value = str(customerDict[key].usesCinder)
    newSheet.cell(row=rowIndex, column=4).value = str(customerDict[key].usesManila)
    newSheet.cell(row=rowIndex, column=5).value = str(customerDict[key].usesONTAP)
    newSheet.cell(row=rowIndex, column=6).value = str(customerDict[key].usesESeries)
    rowIndex += 1

try:
    outputExcel.save('Output.xlsx')
    print "Success!"
except:
    print "Unable to save file. Please ensure you have sufficient disk space " \
          "and permissions."




